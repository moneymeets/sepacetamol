from dataclasses import dataclass
from enum import Enum, unique

from django.core.exceptions import SuspiciousOperation
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render
from django.utils import timezone
from django.utils.encoding import smart_str
from openpyxl import load_workbook
from schwifty import IBAN
from sepaxml import SepaTransfer


@dataclass
class Originator:
    name: str
    iban: str
    bic: str


@dataclass
class Transaction:
    name: str
    iban: str
    bic: str
    amount: float
    purpose: str
    reference: str


def parse_iban(iban: str) -> IBAN:
    try:
        return IBAN(iban)
    except ValueError as e:
        raise SuspiciousOperation(e) from e


def index(request):
    source_file = request.FILES["source-file"] if request.method == "POST" else None

    originator = None
    transactions = []

    if request.method == "POST":
        worksheet = load_workbook(source_file).active

        (name, iban, *_), *_ = worksheet.iter_rows(min_row=2, max_row=2, values_only=True)
        originator = Originator(name=name, iban=parse_iban(iban).formatted, bic=parse_iban(iban).bic)

        for row in worksheet.iter_rows(min_row=4, values_only=True):
            if not any(row):
                continue

            name, iban, amount, purpose, reference = row

            iban = parse_iban(iban)
            bic = str(iban.bic) if iban.country_code == "DE" else ""

            transactions.append(
                Transaction(
                    name=name,
                    iban=iban.formatted,
                    bic=bic,
                    amount=amount,
                    purpose=purpose,
                    reference=reference if reference is not None else "",
                ),
            )

    source_filename = source_file.name if source_file is not None else None
    target_filename = source_filename.replace(".xlsx", ".xml") if source_filename is not None else None

    return render(
        request,
        "index.html",
        {
            "source_filename": source_filename,
            "target_filename": target_filename,
            "originator": originator,
            "transactions": transactions,
            "grand_total": sum(transaction.amount for transaction in transactions),
        },
    )


@unique
class BatchBooking(Enum):
    TRUE = "true"
    FALSE = "false"
    SINGLE = "single"


def generate(request):
    if request.method != "POST":
        return HttpResponseBadRequest

    target_filename = request.POST["target-filename"]
    batch_booking = BatchBooking(request.POST["batch-booking"])

    originator = Originator(
        *[request.POST[field].strip() for field in ("originator-name", "originator-iban", "originator-bic")],
    )

    iban = parse_iban(originator.iban)
    assert iban.country_code == "DE", "only German originator IBANs are supported"

    sepa = SepaTransfer(
        {
            "name": originator.name,
            "IBAN": str(iban),
            "BIC": originator.bic,
            "batch": batch_booking != BatchBooking.SINGLE,
            "currency": "EUR",
        },
        clean=True,
    )

    for name, iban, bic, amount, purpose, reference in zip(
        *(
            (field.strip() for field in request.POST.getlist(item))
            for item in (
                "transaction-name",
                "transaction-iban",
                "transaction-bic",
                "transaction-amount",
                "transaction-purpose",
                "transaction-reference",
            )
        ),
    ):
        sepa.add_payment(
            {
                "name": name,
                "IBAN": str(parse_iban(iban)),
                "BIC": bic,
                "amount": int(float(amount) * 100),  # cents
                "description": purpose,
                "execution_date": timezone.now().date(),
                "endtoend_id": reference if reference != "" else "NOTPROVIDED",
            },
        )

    contents = sepa.export(validate=True)

    if batch_booking == BatchBooking.FALSE:
        contents = contents.replace(b"<BtchBookg>true", b"<BtchBookg>false")

    response = HttpResponse(content_type="application/xml")
    response["Content-Disposition"] = "attachment; filename=%s" % smart_str(target_filename)
    response.write(contents)

    return response
