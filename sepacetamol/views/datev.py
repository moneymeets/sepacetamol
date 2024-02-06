import calendar
import csv
import re
from datetime import date, datetime
from io import StringIO
from typing import Annotated, Literal, Optional
from zoneinfo import ZoneInfo

from django.contrib import messages as message
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render
from django.utils.encoding import smart_str
from openpyxl.reader.excel import load_workbook
from pydantic import BaseModel, ConfigDict, Field, field_validator


def float_to_german(value: float) -> str:
    return f"{value:.2f}".replace(".", ",")


def unquote_empty_csv_strings(value: str) -> str:
    while ';"";' in value:
        value = value.replace(';"";', ";;")
    result = value.removesuffix(';""')
    return result if result == value else result + ";"


def date_to_datev(d: date) -> str:
    return d.strftime("%Y%m%d")


DEFAULT_CONFIG = ConfigDict(
    populate_by_name=True,
    validate_default=True,
    validate_assignment=True,
    frozen=True,
)


PositiveInt = Annotated[int, Field(gt=0)]


class DatevSettings(BaseModel):
    consultant_number: PositiveInt
    client_numer: PositiveInt

    model_config = DEFAULT_CONFIG


class DatevModel(BaseModel):
    def model_dump_csv(self) -> str:
        output = StringIO()
        datev_writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC, delimiter=";")
        datev_writer.writerow(self.model_dump().values())
        return "\n".join(unquote_empty_csv_strings(line) for line in output.getvalue().splitlines())

    model_config = ConfigDict(
        populate_by_name=True,
        validate_default=True,
        validate_assignment=True,
        frozen=True,
    )


class DatevHeader(DatevModel):
    flag: Literal["EXTF", "DTVF"]
    version_numer: Literal[700] = 700
    format_category: Literal[16, 20, 21, 46, 48, 65]
    format_name: Literal[
        "Buchungsstapel",
        "Wiederkehrende Buchungen",
        "Debitoren/Kreditoren",
        "Sachkontenbeschriftungen",
        "Zahlungsbedingungen",
        "Diverse Adressen",
    ]
    format_version: Literal[2, 4, 5, 9, 12]
    created_on: Optional[str] = Field(
        default=None,
        pattern=(
            r"^20[0-9]{2}(0[1-9]|1[0-2])(0[1-9]|[1-2][0-9]|3[0-1])"
            r"(2[0-3]|[01][0-9])[0-5][0-9][0-5][0-9][0-9][0-9][0-9]$"
        ),
    )
    reserved_07: str = Field(default="", max_length=0)
    reserved_08: Optional[str] = Field(default=None, pattern=r"^\w{1,2}$")
    reserved_09: Optional[str] = Field(default=None, pattern=r"^\w{1,25}$")
    reserved_10: Optional[str] = Field(default=None, pattern=r"^\w{1,25}$")
    consultant_number: int = Field(ge=1001, le=9999999)
    client_numer: int = Field(ge=1, le=99999)
    business_year_start: int
    gl_account_length: int = Field(ge=4, le=8)
    date_from: int
    date_till: int
    designation: str = Field(pattern=r"^[\w.\-/ ]{0,30}$")
    initials: Optional[str] = Field(default=None, pattern=r"^([A-Z]{2}){1,2}$")
    record_type: Optional[Literal[1, 2]] = None
    accounting_reason: Optional[Literal[0, 30, 40, 50, 64]] = None
    locking: Literal[0, 1]
    currency_code: str = Field(default="EUR", pattern=r"^[A-Z]{3}$")
    reserved_23: str = Field(default="", max_length=0)
    derivatives_flag: str = Field(default="", max_length=0)
    reserved_25: str = Field(default="", max_length=0)
    reserved_26: str = Field(default="", max_length=0)
    gl_chart_of_accounts: str = Field(pattern=r"^(\d{2}){0,2}$")
    industry_solution_id: Optional[int] = Field(default=None, ge=0, le=9999)
    reserved_29: str = Field(default="", max_length=0)
    reserved_30: str = Field(default="", max_length=0)
    application_information: str = Field(default="", max_length=16)

    @field_validator("business_year_start", "date_from", "date_till")
    def date_pattern(cls, v):
        if re.match(r"^20[0-9]{2}(0[1-9]|1[0-2])(0[1-9]|[1-2][0-9]|3[0-1])$", str(v)) is None:
            raise ValueError("must match date pattern")
        return v


class DatevBooking(DatevModel):
    umsatz: str = Field(
        alias="Umsatz",
        pattern=r"^\d{1,10}[,]\d{2}$",
        description="Umsatz/Betrag für den Datensatz, z.B.: 1234567890,12 .Betrag muss positiv sein.",
    )
    soll_haben_kz: Literal["S", "H"] = Field(
        alias="Soll-/Haben-Kennzeichen",
        description="Soll-/Haben-Kennzeichnung bezieht sich auf das Feld #7 Konto",
    )
    wkz_umsatz: Optional[str] = Field(
        alias="WKZ Umsatz",
        default=None,
        pattern=r"^[A-Z]{3}$",
        description="ISO-Code der Währung",
    )
    kurs: Optional[str] = Field(
        alias="Kurs",
        default=None,
        pattern=r"^[1-9]\d{0,3}[,]\d{2,6}$",
        description="Wenn Umsatz in Fremdwährung bei #1 angegeben wird #4, #5 und #6 sind zu übergeben, z.B.: 12,34",
    )
    basisumsatz: Optional[str] = Field(
        alias="Basisumsatz",
        default=None,
        pattern=r"^\d{1,10}[,]\d{2}$",
        description="Siehe #4, z.B.: 1234567890,12",
    )
    wkz_basisumsatz: Optional[str] = Field(
        alias="WKZ Basisumsatz",
        default=None,
        pattern=r"^[A-Z]{3}$",
        description="Siehe #4",
    )
    konto: int = Field(alias="Konto", description="Sach- oder Personenkonto, z.B. 8400")
    gegenkonto: int = Field(alias="Gegenkonto (ohne BU-Schlüssel)", description="Sach- oder Personenkonto, z.B. 70000")
    bu_schluessel: Optional[int] = Field(
        alias="BU-Schlüssel",
        default=None,
        ge=1000,
        le=9999,
        description="Steuerungskennzeichen zur Abbildung verschiedener Funktionen/Sachverhalte",
    )
    belegdatum: str = Field(
        alias="Belegdatum",
        pattern=r"^\d{4}$",
        description="Format: TTMM, z.B. 0105, Das Jahr wird immer aus dem Feld 13 des Headers ermittelt",
    )
    belegfeld_1: Optional[str] = Field(
        alias="Belegfeld 1",
        default=None,
        pattern=r"^[\w$%\-/]{0,36}$",
        description="Rechnungs-/Belegnummer, wird als 'Schlüssel' für den Ausgleich offener Rechnungen genutzt",
    )
    belegfeld_2: Optional[str] = Field(
        alias="Belegfeld 2",
        default=None,
        pattern=r"^[\w$%\-/]{0,12}$",
        description="Mehrere Funktionen",
    )
    skonto: Optional[str] = Field(
        alias="Skonto",
        default=None,
        pattern=r"^([1-9]\d{0,7}[,]\d{2})$",
        description="Skontobetrag, z.B. 3,71 - nur bei Zahlungsbuchungen zulässig",
    )
    buchungstext: str = Field(
        alias="Buchungstext",
        max_length=60,
        description="0-60 Zeichen",
    )
    postensperre: Optional[Literal[0, 1]] = Field(
        alias="Postensperre",
        default=None,
        description="Mahn- oder Zahlsperre, 0 = keine Sperre (default), 1 = Sperre",
    )

    def model_dump_csv_header(self) -> str:
        output = StringIO()
        datev_writer = csv.DictWriter(
            output,
            fieldnames=self.model_dump(by_alias=True).keys(),
            quoting=csv.QUOTE_NONNUMERIC,
            delimiter=";",
        )
        datev_writer.writeheader()
        return output.getvalue().strip()


def get_datev_booking_from_personio(row: dict) -> tuple[date, DatevBooking]:
    datum = row["Datum"]
    umsatz = row["Umsatz"]
    sh = row["S/H"]
    gegenkonto = row["Gegenkonto"]
    konto = row["Konto"]
    belegfeld_1 = row.get("Belegfeld 1") or row.get("Beleg Feld 1", "")
    buchungstext = row["Buchungstext"]

    # Personio hacks to support different formats
    datum = (
        datetime.strptime(datum, "%d.%m.%Y").replace(tzinfo=ZoneInfo("Europe/Berlin")).date()
        if isinstance(datum, str)
        else datum.date()
    )

    if isinstance(belegfeld_1, float):
        belegfeld_1 = str(int(belegfeld_1))

    buchungstext = (buchungstext[:57] + "...") if len(buchungstext) > 60 else buchungstext

    soll_haben = "S" if not sh else sh

    if umsatz < 0:
        soll_haben = "H" if soll_haben == "S" else "S"

    return (
        datum,
        DatevBooking(
            umsatz=float_to_german(abs(umsatz)),
            soll_haben_kz=soll_haben,
            konto=konto,
            gegenkonto=gegenkonto,
            belegdatum=datum.strftime("%d%m"),
            belegfeld_1=belegfeld_1,
            buchungstext=buchungstext,
        ),
    )


def convert_personio_to_datev(request) -> HttpResponse:
    datev_settings = DatevSettings(
        consultant_number=request.POST["consultant-number"],
        client_numer=request.POST["client-number"],
    )

    try:
        worksheet = load_workbook(request.FILES["personio-file"]).active
    except Exception as e:
        raise ValueError("Personio file could not be loaded, please check the format") from e

    non_empty_rows = tuple(
        tuple(value.strip() if isinstance(value, str) else value for value in row)
        for row in worksheet.iter_rows(values_only=True)
        if any(row)
    )

    bookings = tuple(get_datev_booking_from_personio(dict(zip(non_empty_rows[0], row))) for row in non_empty_rows[1:])

    (datum, first_booking), *_ = bookings

    header = DatevHeader(
        flag="EXTF",
        format_category=21,
        format_name="Buchungsstapel",
        format_version=9,
        consultant_number=datev_settings.consultant_number,
        client_numer=datev_settings.client_numer,
        business_year_start=date_to_datev(datum.replace(month=1, day=1)),
        gl_account_length=4,
        date_from=date_to_datev(datum.replace(day=1)),
        date_till=date_to_datev(datum.replace(day=calendar.monthrange(datum.year, datum.month)[1])),
        designation=f"Lohnbuchungen {datum.strftime('%Y-%m')}",
        currency_code="EUR",
        locking=0,
        gl_chart_of_accounts="03",
    )

    target_filename = f"EXTF_Personio-{datum.strftime('%Y-%m')}.csv"

    contents = "\r\n".join(
        (
            header.model_dump_csv(),
            first_booking.model_dump_csv_header(),
            *(booking.model_dump_csv() for _, booking in bookings),
        ),
    ).encode("cp1252")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=%s" % smart_str(target_filename)
    response.write(contents)

    return response


def index(request):
    if request.method == "GET":
        return render(request, "datev.html")
    elif request.method == "POST":
        try:
            response = convert_personio_to_datev(request)
        except Exception as e:
            message.error(request, e)
            return render(request, "datev.html")
        else:
            return response
    else:
        return HttpResponseBadRequest
